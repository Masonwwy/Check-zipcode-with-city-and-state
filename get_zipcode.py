from openpyxl import load_workbook
import requests
import json
from threading import Thread


def get_zip(city, state, index):
    requests.packages.urllib3.disable_warnings()
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/67.0.3396.87 Safari/537.36'
    }
    try:
        url = 'https://m.usps.com/m/QuickZipAction?mode=0&tCity=' + city + '&sState=' + state + '&jsonInd=Y'
        session = requests.session()
        content = session.post(url, headers=headers, verify=False)
        a = json.loads(content.text)

        global result
        result = []
        result.append(str(index))
        result.append(a["addresses"][0]['zip'])

        final.append(result)
        result = []
        return a["addresses"][0]['zip']
    except:
        return "Error with "+str(city)+' - '+str(state)
        pass


def main():
    wb = load_workbook('result.xlsx')
    st = wb['Sheet1']
    city_state = []
    locs = []

    for i in range(1, int(st.max_row+1)):
        city_state.append(st.cell(i, 1).value)
        city_state.append(st.cell(i, 2).value)
        locs.append(city_state)
        city_state = []

    ts = []
    for loc in locs:
        t = Thread(target=get_zip, args=[loc[0], loc[1], locs.index(loc)])
        t.start()
        ts.append(t)
    for t in ts:
        t.join()

    for item in final:
        try:
            st.cell(int(item[0])+1, 3).value = item[1]
            wb.save("result.xlsx")
        except:
            pass


if __name__ == '__main__':
    final = []
    main()
    print('Done')
